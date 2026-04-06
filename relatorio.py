import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import os

def estilizar_cabecalho(ws, colunas, fill, fonte, borda):
    """APLICA ESTILO AO CABEÇALHO DE UMA ABA."""
    for col_idx, nome_coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx, value=nome_coluna)
        celula.fill = fill
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

def estilizar_dados(ws, df, fill_par, fill_impar, borda):
    """Escreve e estiliza os dados de um DataFrame numa aba."""
    for row_idx, (_, linha) in enumerate(df.iterrows(), start=2):
        for col_idx, valor in enumerate(linha, start=1):
            celula = ws.cell(row=row_idx, column=col_idx, value=valor)
            celula.border = borda
            celula.alignment = Alignment(horizontal="center")
            celula.fill = fill_par if row_idx % 2 == 0 else fill_impar

def ajustar_colunas(ws, colunas):
    """Ajusta largura das colunas automaticamente."""
    for col_idx, nome_coluna in enumerate(colunas, start=1):
        letra = get_column_letter(col_idx)
        largura = max(len(str(nome_coluna)), 15) + 4
        ws.column_dimensions[letra].width = largura

def gerar_aba(wb, titulo, df, fill_cabecalho, fonte_cabecalho, fill_par, fill_impar, borda, primeira=False):
    """
    Cria uma aba no workbook com os dados formatados.

    Parâmetros:
        wb            — Workbook do openpyxl
        titulo        — Nome da aba
        df            — DataFrame com os dados
        primeira      — True se for a primeira aba (usa wb.active)
    """
    ws = wb.active if primeira else wb.create_sheet(title=titulo)
    ws.title = titulo
    colunas = list(df.columns)
    estilizar_cabecalho(ws, colunas, fill_cabecalho, fonte_cabecalho, borda)
    estilizar_dados(ws, df, fill_par, fill_impar, borda)
    ajustar_colunas(ws, colunas)

def gerar_relatorio(custo, qualidade, especializacoes, nome_saida):
    """
    Gera um arquivo Excel com 3 abas formatadas.

    Parâmetros:
        custo           — DataFrame de custo de contratação
        qualidade       — DataFrame de qualidade de vida
        especializacoes — DataFrame de especializações de IA
        nome_saida      — caminho completo do arquivo a ser criado
    """
    caminho = nome_saida

    # --- ESTILO DO CABEÇALHO ------
    #PatternFill DEFINE A COR DE FUNDO DA CÉLULA
    #fgColor É O CÓDIGO HEXADECIMAL DA COR(SEM O #)

    fill_cabecalho = PatternFill(
        start_color="25047A",
        end_color="4B0082",
        fill_type="solid"
    )
    fonte_cabecalho = Font(
        bold=True,
        color="FF8C00",
        size=12
    )

    #---- ESTILO DAS LINHAS ALTERNADAS------
    fill_impar = PatternFill(
        start_color="C0C0C0",
        end_color="C0C0C0",
        fill_type="solid"
    )

    fill_par = PatternFill(
        start_color="7B68EE",
        end_color="7B68EE",
        fill_type="solid"  
    )

    #--- ESTILO DE BORDA FINA -----
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    wb = openpyxl.Workbook()

    # -- Aba 1: Custo de Contratação -------------
    ws1 = wb.active
    ws1.title = "Custo de Constratação"
    colunas1 = list(custo.columns)
    estilizar_cabecalho(ws1, colunas1, fill_cabecalho, fonte_cabecalho, borda)
    estilizar_dados(ws1, custo, fill_par, fill_impar, borda)
    ajustar_colunas(ws1, colunas1)

    # ── Aba 2: Qualidade de Vida ──────────────────────────
    ws2 = wb.create_sheet(title="Qualidade de Vida")
    colunas2 = list(qualidade.columns)
    estilizar_cabecalho(ws2, colunas2, fill_cabecalho, fonte_cabecalho, borda)
    estilizar_dados(ws2, qualidade, fill_par, fill_impar, borda)
    ajustar_colunas(ws2, colunas2)

    # ── Aba 3: Especializações de IA ──────────────────────
    ws3 = wb.create_sheet(title="Especializações de IA")
    colunas3 = list(especializacoes.columns)
    estilizar_cabecalho(ws3, colunas3, fill_cabecalho, fonte_cabecalho, borda)
    estilizar_dados(ws3, especializacoes, fill_par, fill_impar, borda)
    ajustar_colunas(ws3, colunas3)

        #----- SALVA O ARQUIVO ------------------------
    wb.save(caminho)
    logging.info(f"Relatório salvo em: {caminho}")
    print(f"Relatório salvo em: {caminho}")

    return
